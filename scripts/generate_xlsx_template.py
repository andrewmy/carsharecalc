from __future__ import annotations

from pathlib import Path
import csv

from openpyxl import Workbook


def read_tsv(path: Path) -> list[list[str]]:
    rows: list[list[str]] = []
    with path.open('r', encoding='utf-8') as f:
        for row in csv.reader(f, delimiter='\t'):
            rows.append(row)
    return rows


def write_sheet(ws, rows: list[list[str]]) -> None:
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            v = '' if value is None else str(value)
            ws.cell(row=r_idx, column=c_idx, value=v)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    templates = root / 'templates' / 'sheets'

    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--variant', choices=['auto-night', 'manual-night'], default='auto-night')
    parser.add_argument('--embed-sheets-formulas', action='store_true', help='Embed Google Sheets formulas into Results/Details (may trigger Excel repair warnings)')
    parser.add_argument('--out', default=None)
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    inputs_file = 'Inputs.tsv' if args.variant == 'auto-night' else 'Inputs.manual_night.tsv'
    details_file = 'Details.tsv' if args.embed_sheets_formulas else 'Details.blank.tsv'
    tab_files = [
        ('Inputs', inputs_file),
        ('Providers', 'Providers.tsv'),
        ('Vehicles', 'Vehicles.tsv'),
        ('Options', 'Options.tsv'),
        ('Results', 'Results.tsv'),
        ('Details', details_file),
    ]

    for tab_name, filename in tab_files:
        ws = wb.create_sheet(title=tab_name)
        rows = read_tsv(templates / filename)
        write_sheet(ws, rows)

    setup_md = (templates / 'SETUP.md').read_text(encoding='utf-8')

    def extract_formula(cell_ref: str) -> str:
        # Expect a bullet like: - `Results!AF2` (total_eur):
        #   - `=ARRAYFORMULA(...)`
        idx = setup_md.find(cell_ref)
        if idx == -1:
            raise RuntimeError(f'Missing {cell_ref} in SETUP.md')
        chunk = setup_md[idx:]
        start = chunk.find('`=')
        if start == -1:
            raise RuntimeError(f'No formula found for {cell_ref}')
        end = chunk.find('`', start + 1)
        return chunk[start + 1 : end]

    if args.embed_sheets_formulas:
        # Add the Google Sheets formulas into Results.
        # Excel may remove/repair these (ARRAYFORMULA etc), but Google Sheets will evaluate after upload.
        ws = wb['Results']
        for ref, addr in [
            ('`Results!A2`', 'A2'),
            ('`Results!AH2`', 'AH2'),
            ('`Results!AI2`', 'AI2'),
            ('`Results!AJ2`', 'AJ2'),
            ('`Results!AK2`', 'AK2'),
            ('`Results!AL2`', 'AL2'),
            ('`Results!AG2`', 'AG2'),
            ('`Results!AF2`', 'AF2'),
        ]:
            ws[addr].value = extract_formula(ref)

    if args.out:
        out = Path(args.out)
        if not out.is_absolute():
            out = root / out
    else:
        suffix = '' if not args.embed_sheets_formulas else '-gs'
        out_name = (
            f'carcalc-template{suffix}.xlsx'
            if args.variant == 'auto-night'
            else f'carcalc-template-manual-night{suffix}.xlsx'
        )
        out = root / out_name
    wb.save(out)
    print(out)


if __name__ == '__main__':
    main()
